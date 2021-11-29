import openpyxl
import speech_recognition as sr
from googlesearch import search
import pyttsx3
import pywhatkit
import datetime
import wikipedia
import openpyxl
import tkinter as tk
from tkinter import PhotoImage

listener = sr.Recognizer()
engine = pyttsx3.init()
voices = engine.getProperty('voices')
engine.setProperty('voice', voices[1].id)


def talk(text):
    engine.say(text)
    engine.runAndWait()


def take_command():
    try:
        with sr.Microphone() as source:
            print('listening...')
            voice = listener.record(source, duration=5)
            command = listener.recognize_google(voice)
            command = command.lower()
            if 'john' in command:
                command = command.replace('john', '')
                print(command)
    except:
        return ""
    return command


def attendence(key):
    wb = openpyxl.load_workbook("att.xlsx", data_only=True)
    sheets = wb.sheetnames
    count = 0
    sh1 = wb['Sheet1']
    column = sh1['A']
    col_li = [column[x].value for x in range(len(column))]
    for x in range(len(column)):
        # pass
        if column[x].value.lower() in key:
            row = sh1[x + 1]
            row_li = [row[x].value for x in range(len(row))]
            for x in range(len(row_li)):
                if row_li[x] == 'p':
                    count += 1
            return str(count)
        else:
            pass
    #dat = "The attendence of " + key + "is " + data
    return "null"


def relay():
    wb = openpyxl.load_workbook("att.xlsx", data_only=True)
    sh1 = wb['Sheet1']
    column = sh1['A']
    col_li = [column[x].value for x in range(len(column))]
    wb = openpyxl.load_workbook("att.xlsx")
    sh1 = wb['Sheet1']
    sh1.insert_cols(3)
    wb.save("att.xlsx")
    wb = openpyxl.load_workbook("att.xlsx")
    sh1 = wb['Sheet1']
    dat = datetime.date.today()
    ws = wb.active
    ws['C1'] = str(dat)
    for i in range(2, len(column)+1):
        j = "A" + str(i)
        talk(sh1[j].value)
        with sr.Microphone() as source:
            voice = listener.record(source, duration=5)
            command = listener.recognize_google(voice)
            command = command.lower()
            if 'present' in command or 'yes' in command:
                cel = "C" + str(i)
                ws[cel] = "p"
            else:
                cel = "C" + str(i)
                ws[cel] = "a"
    wb.save("att.xlsx")
    talk("Attendance has been recorded")

def Createfile():
    dat = str(datetime.date.today())
    fname = "C:\\Users\\preet\\PycharmProjects\\assignment\\mylogs\\" + dat + ".txt"
    file1 = open(fname, "w")
    talk("What is the log entry")
    with sr.Microphone() as source:
        voice = listener.record(source, duration=12)
        entry = listener.recognize_google(voice)
        entry = entry.lower()
   # L = ["This is Delhi \n", "This is Paris \n", "This is London"]
        file1.writelines(entry)
    talk("Log has been created successfully")
    file1.close()

def percent(key):
    wb = openpyxl.load_workbook("book3.xlsx", data_only=True)
    sheets = wb.sheetnames
    count = 0
    sh1 = wb['Sheet1']
    # dat = sh1['B2'].value
    column = sh1['1']
    col_li = [column[x].value for x in range(len(column))]
    print(col_li)
    tot = len(column)
    tot -= 2
    #row_li = []
    column = sh1['A']
    col_li = [column[x].value for x in range(len(column))]
    for x in range(len(column)):
        print(x)
        if column[x].value.lower() in key:
            row = sh1[x + 1]
            row_li = [row[x].value for x in range(len(row))]
            print(row_li)
            for x in range(len(row_li)):
                if row_li[x] == 'p':
                    count += 1
    print(count)
    perc = (count / tot) * 100
    perc = round(perc,2)
    print(perc)
    per = str(perc)
    talk("Attendance of " + key + "is " + per + "Percent")


def Updatefile():
    dat = str(datetime.date.today())
    fname = "C:\\Users\\preet\\PycharmProjects\\assignment\\mylogs\\" + dat + ".txt"
    file1 = open(fname, "a")
    talk("What is the log entry")
    with sr.Microphone() as source:
        voice = listener.record(source, duration=12)
        entry = listener.recognize_google(voice)
        entry = entry.lower()
        # L = ["This is Delhi \n", "This is Paris \n", "This is London"]
        file1.writelines(entry)
    talk("Log has been updated successfully")
    file1.close()

def dictres(key):
    wb = openpyxl.load_workbook("marks.xlsx")
    sh1 = wb['Sheet1']
    column = sh1['A']
    col_li = [column[x].value for x in range(len(column))]
    print(key)
    if "test1" in key or "test 1" in key:
        column2 = sh1['C']
        col_li2 = [column2[x].value for x in range(len(column2))]
        for x in range(1,len(column2)):
            talk(str(col_li[x])+" got "+str(col_li2[x]))
    elif "test2" in key or "test 2" in key:
        column2 = sh1['D']
        col_li2 = [column2[x].value for x in range(len(column2))]
        for x in range(1,len(column2)):
            talk(str(col_li[x])+" got "+str(col_li2[x]))
    elif "test3" in key or "test 3" in key:
        column2 = sh1['E']
        col_li2 = [column2[x].value for x in range(len(column2))]
        for x in range(1,len(column2)):
            talk(str(col_li[x])+" got "+str(col_li2[x]))
    elif "quiz1" in key or "quiz 1" in key:
        column2 = sh1['F']
        col_li2 = [column2[x].value for x in range(len(column2))]
        for x in range(1,len(column2)):
            talk(str(col_li[x])+" got "+str(col_li2[x]))
    elif "quiz2" in key or "quiz 2" in key:
        column2 = sh1['G']
        col_li2 = [column2[x].value for x in range(len(column2))]
        for x in range(1,len(column2)):
            talk(str(col_li[x])+" got "+str(col_li2[x]))
    elif "assignment1" in key or "assignment 1" in key:
        column2 = sh1['H']
        col_li2 = [column2[x].value for x in range(len(column2))]
        for x in range(1,len(column2)):
            talk(str(col_li[x])+" got "+str(col_li2[x]))
    elif "assignment1" in key or "assignment 2" in key:
        column2 = sh1['I']
        col_li2 = [column2[x].value for x in range(len(column2))]
        for x in range(1,len(column2)):
            talk(str(col_li[x])+" got "+str(col_li2[x]))
    elif "final" in key or "final" in key:
        column2 = sh1['J']
        col_li2 = [column2[x].value for x in range(len(column2))]
        for x in range(1,len(column2)):
            talk(str(col_li[x])+" got "+str(col_li2[x]))
    elif "total" in key or "total" in key:
        column2 = sh1['B']
        col_li2 = [column2[x].value for x in range(len(column2))]
        for x in range(1,len(column2)):
            talk(str(col_li[x])+" got "+str(col_li2[x]))


def entermarks(key):
    wb = openpyxl.load_workbook("marks.xlsx")
    sh1 = wb['Sheet1']
    column = sh1['A']
    sub = ""
    usr = 9999
    col_li = [column[x].value for x in range(len(column))]
    for x in range(len(column)):
        # pass
        if column[x].value.lower() in key:
            usr = x + 1
            talk("Which marks would you like to update?")
            with sr.Microphone() as source:
                voice = listener.record(source, duration=5)
                entry = listener.recognize_google(voice)
                sub = entry.lower()
            talk("marks obtained?")
            with sr.Microphone() as source:
                voice = listener.record(source, duration=5)
                entry = listener.recognize_google(voice)
                mark = entry.lower()
            print(sub)
            ws = wb.active
            if sub == "test1" or sub == "test 1":
                cell = "C" + str(usr)
                ws[cell] = mark
                print(cell)
            elif sub == "test2" or sub == "test 2":
                cell = "D" + str(usr)
                ws[cell] = mark
                print(cell)
            elif sub == "test3" or sub == "test 3":
                cell = "D" + str(usr)
                ws[cell] = str(mark)
            elif sub == "assignment1" or sub == "assignment 1":
                cell = "D" + str(usr)
                ws[cell] = str(mark)
            elif sub == "assignment2" or sub == "assignment 2":
                cell = "D" + str(usr)
                ws[cell] = str(mark)
            elif sub == "final" or sub == "final":
                cell = "D" + str(usr)
                ws[cell] = str(mark)

            wb.save("marks.xlsx")
            talk("marks have been updated")


def usn_search():
    wb = openpyxl.load_workbook("studdb.xlsx")
    sh1 = wb['Sheet1']
    # key = 'Preetam'
    talk("Please tell the group")
    with sr.Microphone() as source:
        voice = listener.record(source, duration=5)
        entry = listener.recognize_google(voice)
        inp = entry.lower()
        key = inp + " "
    if 'z' in key:
        key = "1RZ19MCA"
    elif 'd' in key:
        key = "1RD19MCA"
    elif 'v' in key:
        key = "1RV19MCA"
    talk("What is the number?")
    with sr.Microphone() as source:
        voice = listener.record(source, duration=5)
        entry = listener.recognize_google(voice)
        inp = entry.lower()
        num = inp + " "
    key += num
    print(key)
    column = sh1['B']
    talk("What would you like to fetch?")
    with sr.Microphone() as source:
        voice = listener.record(source, duration=5)
        entry = listener.recognize_google(voice)
        inp = entry.lower()
        ch = inp + " "
    print(ch)
    col_li = [column[x].value for x in range(len(column))]
    print(col_li)
    for x in range(len(col_li)):
        if col_li[x] in key:
            j = x + 1
            row = sh1[j]
            # row_li = [row[y].value for y in range(len(row))]
            # print(row_li)
            if 'name' in ch or 'us' in ch:
                z = "A" + str(j)
                talk("Name is " + sh1[z].value)
            if 'number' in ch:
                z = "C" + str(j)
                talk("number is " + str(sh1[z].value))
            if 'councillor' in ch or 'counselor' in ch:
                z = "D" + str(j)
                talk("councillor is " + sh1[z].value)
            if 'batch' in ch or 'bach' in ch:
                z = "E" + str(j)
                talk("batch is " + sh1[z].value)


def name_search():
    wb = openpyxl.load_workbook("studdb.xlsx")
    sh1 = wb['Sheet1']
    #key = 'Preetam'
    talk("Please tell the name")
    with sr.Microphone() as source:
        voice = listener.record(source, duration=5)
        entry = listener.recognize_google(voice)
        inp = entry.lower()
        key = inp + " "
    print(key)
    column = sh1['A']
    talk("What would you like to fetch?")
    with sr.Microphone() as source:
        voice = listener.record(source, duration=5)
        entry = listener.recognize_google(voice)
        inp = entry.lower()
        ch = inp + " "
    print(ch)
    col_li = [column[x].value for x in range(len(column))]
    for x in range(len(col_li)):
        if col_li[x] in key:
            j = x + 1
            row = sh1[j]
            # row_li = [row[y].value for y in range(len(row))]
            # print(row_li)
            if 'usn' in ch or 'us' in ch:
                z = "B" + str(j)
                talk("USN is " + sh1[z].value)
            if 'number' in ch:
                z = "C" + str(j)
                talk("number is " + str(sh1[z].value))
            if 'councillor' in ch or 'counselor' in ch:
                z = "D" + str(j)
                talk("councillor is " + sh1[z].value)
            if 'batch' in ch:
                z = "E" + str(j)
                talk("batch is " + sh1[z].value)


def process():
    command = take_command()
    print(command)
    if 'play' in command:
        video = command.replace('play', '')
        talk('playing ' + video)
        pywhatkit.playonyt(video)
        exit(0)
    elif 'time' in command:
        time = datetime.datetime.now().strftime('%I:%M %p')
        talk('Current time is ' + time)
    elif 'who is' in command:
        person = command.replace('who is', '')
        info = wikipedia.summary(person, 1)
        print(info)
        talk(info)
    elif 'what is' in command:
        topic = command.replace('what is', '')
        info = wikipedia.summary(topic, 1)
        print(info)
        talk(info)
    elif 'search' in command:
        qry = command.replace('search', '')
        opt = ""
        for j in search(qry, tld="co.in", num=10, stop=10, pause=2):
            print(j)
            opt += "\n\n" + str(j)
        t.insert(tk.END,opt)
    elif 'attendance of' in command:
        pos = command.find("of")
        key = command[pos:]
        key = key + " "
        dat = attendence(key)
        dat = "Attendance of " + key + "is " + str(dat)
        talk(dat)
    elif 'attendance percent of' in command:
        pos = command.find("of")
        key = command[pos:]
        key = key + " "
        percent(key)
        #dat = "Attendance percent of " + key + "is " + str(dat)
        #talk(dat)
    elif 'conduct attendance' in command:
        relay()
    elif 'create log' in command:
        Createfile()
    elif 'update log' in command:
        Updatefile()
    elif 'dictate marks' in command:
        pos = command.find("of")
        key = command[pos:]
        key = key + " "
        dictres(key)
    elif 'update marks of' in command:
        pos = command.find("of")
        key = command[pos:]
        key = key + " "
        entermarks(key)
    elif 'student details' in command:
        talk("What is your input?")
        with sr.Microphone() as source:
            voice = listener.record(source, duration=5)
            entry = listener.recognize_google(voice)
            inp = entry.lower()
            inp += " "
        print(inp)
        if 'us' in inp or 'sn' in inp or 'u s' in inp:
            usn_search()
        elif 'name' in inp:
            name_search()
    elif 'terminate' in command:
        talk('Thank You')
        exit()
    else:
        talk('Please say the command again.')
        process()


def print_hello():
    talk('Hi i am John. How can i help you?')
    process()

root = tk.Tk()
root.geometry("960x600")

imagetest = PhotoImage(file="mic_img.png")

button_qwer = tk.Button(root, text="Command", image=imagetest, command=print_hello)
button_qwer.pack()
t = tk.Text(root, height=20, width=45)
t.pack()
root.mainloop()
